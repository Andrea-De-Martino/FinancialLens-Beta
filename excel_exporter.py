from openpyxl import Workbook
from openpyxl.styles import (
    Font, Alignment, PatternFill, Border, Side, NamedStyle
)
from openpyxl.utils import get_column_letter
from typing import Dict, Any, Optional
from datetime import datetime


class ProfessionalExcelExporter:
    """
    Excel exporter for financial statement analysis
    """

    def __init__(self):
        """Initialize the Excel exporter."""
        self.color_scheme = self._setup_refined_colors()
        self.styles = self._setup_professional_styles()
        self.workbook = None
        self.current_row = 1

    def export(self, data: Dict[str, Any], excel_output_path: str, 
               reclassification_result: Optional[Dict[str, Any]] = None) -> None:
        """
        Export financial statement data to a formatted Excel workbook.
        
        Args:
            data: Complete financial statement data dictionary
            excel_output_path: Path where the Excel file should be saved
            reclassification_result: Optional reclassification analysis results
        """
        # Extract reclassification data from main data if not provided separately
        if reclassification_result is None:
            reclassification_result = data.get('reclassification', {})
        self.workbook = Workbook()
        
        # Remove default sheet
        self.workbook.remove(self.workbook.active)
        
        # Create sheets in order with updated names
        if reclassification_result and reclassification_result.get('success'):
            self._create_brsf_sheet(reclassification_result)
        
        self._create_balance_sheet(data)
        self._create_income_statement(data)  
        self._create_validation_report(data, reclassification_result)
        
        # Set the first sheet as active
        if self.workbook.worksheets:
            self.workbook.active = self.workbook.worksheets[0]
        
        # Save the workbook
        self.workbook.save(excel_output_path)
        print(f"Excel report created successfully: {excel_output_path}")

    def _setup_refined_colors(self) -> Dict[str, str]:
        """
        Define rbranding.
        
        Returns:
            Dictionary containing color codes
        """
        return {
            # Primary palette
            'text_primary': '2D3748',              # Dark gray for primary text
            'text_secondary': '4A5568',            # Medium gray for secondary text
            'text_muted': '718096',                # Light gray for muted text
            
            # Background colors
            'bg_white': 'FFFFFF',                  # Pure white
            'bg_light': 'F7FAFC',                  # Very light gray background
            'bg_neutral': 'EDF2F7',               # Light neutral background
            'bg_subtle': 'E2E8F0',                # Subtle gray background
            
            # Border colors
            'border_light': 'E2E8F0',             # Light borders
            'border_medium': 'CBD5E0',            # Medium borders
            'border_strong': 'A0AEC0',            # Strong borders
            
            # red (used sparingly as accent)
            'accent_red': 'E53E3E',               # red for highlights only
            'accent_red_light': 'FED7D7',         # Very light red for subtle backgrounds
            'accent_red_soft': 'F56565',          # Softer red for title backgrounds
            
            # Status colors (palette)
            'success': '48BB78',                  # green
            'warning': 'ED8936',                  # orange
            'error': 'E53E3E',                    # red for errors
            
            # Calculated totals (dark)
            'calculated_bg': '4A5568',            # dark gray
            'calculated_text': 'FFFFFF'           # White text on dark
        }

    def _setup_professional_styles(self) -> Dict[str, Any]:
        """
        Create professional cell styles following modern design principles.
        
        Returns:
            Dictionary containing professional style objects
        """
        colors = self.color_scheme
        
        return {
            # Header styles 
            'main_header': NamedStyle(
                name='main_header',
                font=Font(name='Segoe UI', size=18, bold=True, color=colors['bg_white']),
                fill=PatternFill(fill_type='solid', start_color=colors['accent_red_soft']),
                alignment=Alignment(horizontal='center', vertical='center')
            ),
            
            'section_header': NamedStyle(
                name='section_header',
                font=Font(name='Segoe UI', size=12, bold=True, color=colors['text_primary']),
                fill=PatternFill(fill_type='solid', start_color=colors['bg_neutral']),
                alignment=Alignment(horizontal='left', vertical='center'),
                border=Border(
                    left=Side(style='thick', color=colors['accent_red']),
                    bottom=Side(style='thin', color=colors['border_medium'])
                )
            ),
            
            'column_header': NamedStyle(
                name='column_header',
                font=Font(name='Segoe UI', size=11, bold=True, color=colors['text_primary']),
                fill=PatternFill(fill_type='solid', start_color=colors['bg_subtle']),
                alignment=Alignment(horizontal='center', vertical='center'),
                border=Border(
                    bottom=Side(style='medium', color=colors['accent_red_soft']),
                    top=Side(style='thin', color=colors['accent_red_light'])
                )
            ),
            
            # Financial data styles 
            'financial_item_level0': NamedStyle(
                name='financial_item_level0',
                font=Font(name='Segoe UI', size=11, bold=True, color=colors['text_primary']),
                alignment=Alignment(horizontal='left', vertical='center'),
                border=Border(
                    bottom=Side(style='thin', color=colors['border_medium']),
                    left=Side(style='hair', color=colors['accent_red_light'])
                )
            ),
            
            'financial_item_level1': NamedStyle(
                name='financial_item_level1',
                font=Font(name='Segoe UI', size=10, bold=False, color=colors['text_primary']),
                alignment=Alignment(horizontal='left', vertical='center', indent=1),
                border=Border(bottom=Side(style='hair', color=colors['border_light']))
            ),
            
            'financial_item_level2': NamedStyle(
                name='financial_item_level2',
                font=Font(name='Segoe UI', size=10, color=colors['text_secondary']),
                alignment=Alignment(horizontal='left', vertical='center', indent=2),
                border=Border(bottom=Side(style='hair', color=colors['border_light']))
            ),
            
            'financial_item_level3': NamedStyle(
                name='financial_item_level3',
                font=Font(name='Segoe UI', size=9, color=colors['text_muted']),
                alignment=Alignment(horizontal='left', vertical='center', indent=3),
                border=Border(bottom=Side(style='hair', color=colors['border_light']))
            ),
            
            # Value styles 
            'currency_value': NamedStyle(
                name='currency_value',
                font=Font(name='Segoe UI', size=10, color=colors['text_primary']),
                number_format='#,##0.00',
                alignment=Alignment(horizontal='right', vertical='center'),
                border=Border(bottom=Side(style='hair', color=colors['border_light']))
            ),
            
            'currency_total': NamedStyle(
                name='currency_total',
                font=Font(name='Segoe UI', size=11, bold=True, color=colors['text_primary']),
                number_format='#,##0.00',
                alignment=Alignment(horizontal='right', vertical='center'),
                fill=PatternFill(fill_type='solid', start_color=colors['bg_light']),
                border=Border(
                    top=Side(style='thin', color=colors['accent_red_light']),
                    bottom=Side(style='medium', color=colors['accent_red'])
                )
            ),
            
            # Calculated/derived totals - dark styling
            'calculated_item': NamedStyle(
                name='calculated_item',
                font=Font(name='Segoe UI', size=10, bold=True, color=colors['calculated_text']),
                fill=PatternFill(fill_type='solid', start_color=colors['calculated_bg']),
                alignment=Alignment(horizontal='left', vertical='center'),
                border=Border(
                    left=Side(style='thin', color=colors['calculated_bg']),
                    right=Side(style='thin', color=colors['calculated_bg']),
                    top=Side(style='thin', color=colors['calculated_bg']),
                    bottom=Side(style='thin', color=colors['calculated_bg'])
                )
            ),
            
            'calculated_value': NamedStyle(
                name='calculated_value',
                font=Font(name='Segoe UI', size=10, bold=True, color=colors['calculated_text']),
                number_format='#,##0.00',
                fill=PatternFill(fill_type='solid', start_color=colors['calculated_bg']),
                alignment=Alignment(horizontal='right', vertical='center'),
                border=Border(
                    left=Side(style='thin', color=colors['calculated_bg']),
                    right=Side(style='thin', color=colors['calculated_bg']),
                    top=Side(style='thin', color=colors['calculated_bg']),
                    bottom=Side(style='thin', color=colors['calculated_bg'])
                )
            ),
            
            # BRSF specific styles - readable
            'brsf_detail': NamedStyle(
                name='brsf_detail',
                font=Font(name='Segoe UI', size=9, color=colors['text_muted']),
                alignment=Alignment(wrap_text=True, vertical='top', horizontal='left'),
                border=Border(bottom=Side(style='hair', color=colors['border_light']))
            ),
            
            # Metadata styles
            'metadata_label': NamedStyle(
                name='metadata_label',
                font=Font(name='Segoe UI', size=10, bold=True, color=colors['text_primary']),
                alignment=Alignment(horizontal='left', vertical='center')
            ),
            
            'metadata_value': NamedStyle(
                name='metadata_value',
                font=Font(name='Segoe UI', size=10, color=colors['text_secondary']),
                alignment=Alignment(horizontal='left', vertical='center')
            )
        }

    def _create_brsf_sheet(self, reclassification_result: Dict[str, Any]) -> None:
        """
        Create BRSF sheet.
        
        Args:
            reclassification_result: Reclassification analysis results
        """
        ws = self.workbook.create_sheet("BRSF")
        self.current_row = 1
        
        # Main title
        ws.merge_cells(f'A{self.current_row}:I{self.current_row}')
        cell = ws[f'A{self.current_row}']
        cell.value = "BRSF"
        cell.style = self.styles['main_header']
        ws.row_dimensions[self.current_row].height = 35
        self.current_row += 2
        
        reclassified_data = reclassification_result.get('reclassified_data', {})
        reclassified_details = reclassification_result.get('reclassified_details', {})
        
        # Define calculated/derived items that need special formatting
        calculated_items = {
            'Cost of Sales', 'S G & A', 'PRE TAX PROFIT', 'PROFIT AFTER TAX',
            'Quick Assets', 'Other Current', 'Financial/Other fix Ass', 'BAL. SHEET TOT',
            'Total Short Term Debt', 'Sub. Loan/Oth. LT', 'Liabs less net worth',
            'TOT NET WORTH'
        }
        
        # P&L Section
        self._write_brsf_section(ws, "P&L", 'A', reclassified_data.get('P&L', {}), 
                                reclassified_details.get('P&L', {}), calculated_items)
        
        # Assets Section (skip column D for spacing)
        self._write_brsf_section(ws, "ASSETS", 'D', reclassified_data.get('ASSETS', {}),
                                reclassified_details.get('ASSETS', {}), calculated_items)
        
        # Liabilities Section
        self._write_brsf_section(ws, "LIABILITIES", 'G', reclassified_data.get('LIABILITIES', {}),
                                reclassified_details.get('LIABILITIES', {}), calculated_items)
        
        # Hide detail columns for cleaner look
        ws.column_dimensions['C'].hidden = True
        ws.column_dimensions['F'].hidden = True  
        ws.column_dimensions['I'].hidden = True
        
        # Set professional column widths
        for col in ['A', 'D', 'G']:
            ws.column_dimensions[col].width = 30
        for col in ['B', 'E', 'H']:
            ws.column_dimensions[col].width = 18

    def _write_brsf_section(self, ws, section_name: str, start_col: str, 
                           section_data: Dict[str, float], section_details: Dict[str, str],
                           calculated_items: set) -> None:
        """Write a BRSF section"""
        start_row = 3
        
        # Section header - Clean design
        ws.merge_cells(f'{start_col}{start_row}:{chr(ord(start_col)+2)}{start_row}')
        cell = ws[f'{start_col}{start_row}']
        cell.value = section_name
        cell.style = self.styles['section_header']
        
        # Column headers - Professional styling
        ws[f'{start_col}{start_row+1}'].value = "Item"
        ws[f'{start_col}{start_row+1}'].style = self.styles['column_header']
        ws[f'{chr(ord(start_col)+1)}{start_row+1}'].value = "Value"
        ws[f'{chr(ord(start_col)+1)}{start_row+1}'].style = self.styles['column_header']
        ws[f'{chr(ord(start_col)+2)}{start_row+1}'].value = "Details"
        ws[f'{chr(ord(start_col)+2)}{start_row+1}'].style = self.styles['column_header']
        
        # Define section structures
        section_structures = {
            'P&L': [
                'Total Sales', 'Cost of Sales', 'GROSS PROFIT', 'S G & A', 'NET OP. PROFIT',
                'Interest Received', 'Interest Paid', 'Other Income/Expense', 'PRE TAX PROFIT',
                'Taxation', 'PROFIT AFTER TAX', 'Deprecation'
            ],
            'ASSETS': [
                'Cash & Equivalent', 'Trade Debtors', 'Quick Assets', 'Stock & Work in Progr.',
                'Other Current', 'TOT CURR. ASS.', 'Tangible Fxed Assets', 'Intangibles',
                'Goodwill', 'Financial/Other fix Ass', 'TOTAL FIX. ASS.', 'BAL. SHEET TOT'
            ],
            'LIABILITIES': [
                'Overdrafts & STD', 'Current Portion LTD', 'Total Short Term Debt', 'Trade Creditors',
                'Other Current', 'TOT CURR. LIAB.', 'Total Long Term Debt', 'Provisions',
                'Sub. Loan/Oth. LT', 'TOTAL LT LIAB', 'Liabs less net worth', 'Share Holders Funds',
                "Ret'd Earnings/Resv's", 'TOT NET WORTH', 'BAL. SHEET TOT'
            ]
        }
        
        structure = section_structures.get(section_name, [])
        
        for i, item in enumerate(structure, start_row+2):
            # Item name
            item_cell = ws[f'{start_col}{i}']
            item_cell.value = item
            
            # Value
            value_cell = ws[f'{chr(ord(start_col)+1)}{i}']
            value_cell.value = section_data.get(item, 0.0)
            
            # Details (hidden column)
            detail_cell = ws[f'{chr(ord(start_col)+2)}{i}']
            detail_cell.value = section_details.get(item, "")
            detail_cell.style = self.styles['brsf_detail']
            
            # Apply professional formatting for calculated items
            if item in calculated_items:
                item_cell.style = self.styles['calculated_item']
                value_cell.style = self.styles['calculated_value']
            else:
                item_cell.style = self.styles['financial_item_level1']
                value_cell.style = self.styles['currency_value']
            
            # Set appropriate row height
            ws.row_dimensions[i].height = 18

    def _create_balance_sheet(self, data: Dict[str, Any]) -> None:
        """Create professionally formatted balance sheet."""
        ws = self.workbook.create_sheet("A&L")
        self.current_row = 1
        
        # Title 
        ws.merge_cells('A1:C1')
        cell = ws['A1']
        cell.value = "A&L"
        cell.style = self.styles['main_header']
        ws.row_dimensions[1].height = 35
        self.current_row = 3
        
        # Headers
        headers = ["Description", "Value (EUR)", "Notes"]
        for i, header in enumerate(headers, 1):
            cell = ws.cell(row=self.current_row, column=i)
            cell.value = header
            cell.style = self.styles['column_header']
        self.current_row += 1
        
        # Balance sheet data
        sp = data.get('stato_patrimoniale', {})
        
        if 'attivo' in sp:
            self._write_financial_section_enhanced(ws, "ASSETS (ATTIVO)", sp['attivo'], 0)
        
        if 'passivo' in sp:
            self._write_financial_section_enhanced(ws, "LIABILITIES (PASSIVO)", sp['passivo'], 0)
        
        # Column widths
        ws.column_dimensions['A'].width = 70
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 25

    def _create_income_statement(self, data: Dict[str, Any]) -> None:
        """Create income statement."""
        ws = self.workbook.create_sheet("P&L")
        self.current_row = 1
        
        # Title 
        ws.merge_cells('A1:C1')
        cell = ws['A1']
        cell.value = "P&L"
        cell.style = self.styles['main_header']
        ws.row_dimensions[1].height = 35
        self.current_row = 3
        
        # Headers
        headers = ["Description", "Value (EUR)", "Notes"]
        for i, header in enumerate(headers, 1):
            cell = ws.cell(row=self.current_row, column=i)
            cell.value = header
            cell.style = self.styles['column_header']
        self.current_row += 1
        
        # Income statement data
        ce = data.get('conto_economico', {})
        self._write_financial_section_enhanced(ws, "INCOME STATEMENT", ce, 0)
        
        # Column widths
        ws.column_dimensions['A'].width = 70
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 25

    def _create_validation_report(self, data: Dict[str, Any], 
                                 reclassification_result: Optional[Dict[str, Any]] = None) -> None:
        """Create validation report."""
        ws = self.workbook.create_sheet("Validation Report")
        self.current_row = 1
        
        # Title
        ws.merge_cells('A1:B1')
        cell = ws['A1']
        cell.value = "VALIDATION REPORT"
        cell.style = self.styles['main_header']
        ws.row_dimensions[1].height = 35
        self.current_row = 3
        
        # Document metadata
        metadata = data.get('metadata', {})
        
        metadata_items = [
            ("Document", metadata.get('file_name', 'N/A')),
            ("Fiscal Year", metadata.get('anno_bilancio', 'N/A')),
            ("Generated", datetime.now().strftime("%d/%m/%Y %H:%M:%S"))
        ]
        
        for label, value in metadata_items:
            ws[f'A{self.current_row}'].value = label
            ws[f'A{self.current_row}'].style = self.styles['metadata_label']
            
            ws[f'B{self.current_row}'].value = value
            ws[f'B{self.current_row}'].style = self.styles['metadata_value']
            
            self.current_row += 1
        
        self.current_row += 2
        
        # Parser validations
        validations = data.get('validazioni', {})
        if validations:
            ws[f'A{self.current_row}'].value = "PARSER VALIDATIONS"
            ws[f'A{self.current_row}'].style = self.styles['section_header']
            self.current_row += 1
            
            overall_status = validations.get('summary', {}).get('status', 'unknown')
            ws[f'A{self.current_row}'].value = "Overall Status:"
            ws[f'A{self.current_row}'].style = self.styles['metadata_label']
            ws[f'B{self.current_row}'].value = overall_status.upper()
            ws[f'B{self.current_row}'].style = self.styles['metadata_value']
            self.current_row += 2
            
            # Detailed validations
            validation_checks = [
                ("Balance Sheet Equilibrium", validations.get('attivo_vs_passivo')),
                ("Profit Consistency", validations.get('coerenza_utile_esercizio'))
            ]
            
            for check_name, check_data in validation_checks:
                if check_data:
                    ws[f'A{self.current_row}'].value = f"{check_name}:"
                    ws[f'A{self.current_row}'].style = self.styles['metadata_label']
                    
                    status = check_data.get('status', 'unknown')
                    discrepancy = check_data.get('discrepanza', 0)
                    
                    ws[f'B{self.current_row}'].value = f"{status.upper()} (Δ {discrepancy:.2f}€)"
                    ws[f'B{self.current_row}'].style = self.styles['metadata_value']
                    
                    self.current_row += 1
        
        # Reclassification validation
        if reclassification_result and reclassification_result.get('success'):
            self.current_row += 1
            ws[f'A{self.current_row}'].value = "RECLASSIFICATION VALIDATION"
            ws[f'A{self.current_row}'].style = self.styles['section_header']
            self.current_row += 1
            
            balance_validation = reclassification_result.get('balance_sheet_validation', {})
            if balance_validation:
                status = balance_validation.get('status', 'unknown')
                difference = balance_validation.get('difference', 0)
                
                ws[f'A{self.current_row}'].value = "Balance Sheet Equilibrium:"
                ws[f'A{self.current_row}'].style = self.styles['metadata_label']
                ws[f'B{self.current_row}'].value = f"{status.upper()} (Δ {difference:.2f}€)"
                ws[f'B{self.current_row}'].style = self.styles['metadata_value']
                self.current_row += 1
            
            orphan_count = reclassification_result.get('orphan_voices_count', 0)
            ws[f'A{self.current_row}'].value = "Unmapped Voices:"
            ws[f'A{self.current_row}'].style = self.styles['metadata_label']
            ws[f'B{self.current_row}'].value = f"{orphan_count} items"
            ws[f'B{self.current_row}'].style = self.styles['metadata_value']
        
        # Professional column widths
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 40

    def _write_financial_section_enhanced(self, ws, section_title: str, 
                                        section_data: Dict, level: int) -> None:
        """Write financial section with professional hierarchical visualization."""
        if not section_data:
            return
        
        # Section header 
        ws.merge_cells(f'A{self.current_row}:C{self.current_row}')
        cell = ws[f'A{self.current_row}']
        cell.value = section_title
        cell.style = self.styles['section_header']
        self.current_row += 1
        
        # Process items with clear hierarchy
        for item_name, item_data in section_data.items():
            if not isinstance(item_data, dict):
                continue
            
            # Determine appropriate style based on hierarchy level
            if level == 0:
                style_name = 'financial_item_level0'
                value_style = 'currency_total'
            elif level == 1:
                style_name = 'financial_item_level1'
                value_style = 'currency_value'
            elif level == 2:
                style_name = 'financial_item_level2'
                value_style = 'currency_value'
            else:
                style_name = 'financial_item_level3'
                value_style = 'currency_value'
            
            # Special formatting for totals
            if 'totale' in item_name.lower():
                value_style = 'currency_total'
            
            # Item name
            ws[f'A{self.current_row}'].value = item_name
            ws[f'A{self.current_row}'].style = self.styles[style_name]
            
            # Value
            value = item_data.get('valore', 0)
            ws[f'B{self.current_row}'].value = value
            ws[f'B{self.current_row}'].style = self.styles[value_style]
            
            # Notes - Clean indicators
            notes = []
            if item_data.get('from_ni'):
                notes.append("From NI")
            if item_data.get('enriched_from_ni'):
                notes.append("Enriched")
            
            ws[f'C{self.current_row}'].value = " | ".join(notes)
            ws[f'C{self.current_row}'].style = self.styles['metadata_value']
            
            self.current_row += 1
            
            # Process children recursively
            if 'dettaglio' in item_data and item_data['dettaglio']:
                self._write_financial_section_enhanced(ws, "", item_data['dettaglio'], level + 1)
        
        # Add professional spacing after sections
        if level == 0:
            self.current_row += 1


# Legacy compatibility wrapper
class ExcelExporter(ProfessionalExcelExporter):
    """Legacy compatibility wrapper for the Excel exporter."""
    
    def __init__(self):
        super().__init__()
    
    def export(self, data: Dict[str, Any], excel_output_path: str,
               reclassification_result: Optional[Dict[str, Any]] = None) -> None:
        super().export(data, excel_output_path, reclassification_result)